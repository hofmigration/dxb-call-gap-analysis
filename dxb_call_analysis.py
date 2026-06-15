#!/usr/bin/env python3
"""
DXB Team — Call-Attempt Gap Analysis
=====================================
For every DXB-owned contact created in the last 2 months, flags leads that have:
  (A) fewer than MIN_CALLS logged Call activities, OR
  (B) a gap of GAP_DAYS or more days between two consecutive call attempts.

Excludes leads whose lead_stage is "No Answer" or "Schedule Call Back (Call Back)".

Counts ONLY logged Call engagements (not emails / WhatsApp / notes).

Output: dxb_call_gap_analysis.xlsx

HOW TO RUN
----------
1. Set an environment variable with a HubSpot Private App token that has
   crm.objects.contacts.read and crm.objects.calls.read (engagements read):
       export HUBSPOT_TOKEN="pat-na1-xxxxxxxx"
2. pip install requests openpyxl
3. python dxb_call_analysis.py

It is safe to re-run monthly — just change WINDOW_START / WINDOW_END.
"""

import os
import time
import datetime as dt
import requests
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------------
TOKEN = os.environ.get("HUBSPOT_TOKEN")
if not TOKEN:
    raise SystemExit("Set HUBSPOT_TOKEN env var to your HubSpot private app token.")

BASE = "https://api.hubapi.com"
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}

# Analysis window — contacts CREATED in this range (last ~2 months)
WINDOW_START = "2026-04-12"
WINDOW_END   = "2026-06-12"           # exclusive upper bound on createdate

# Call window — calls counted if they occurred between these dates
CALL_START   = "2026-04-12"
CALL_END     = "2026-06-13"

MIN_CALLS = 8          # flag leads with fewer than this many calls
GAP_DAYS  = 10         # flag leads with a gap >= this many days between consecutive calls

# Lead stages to EXCLUDE from the analysis
EXCLUDE_STAGES = {"No Answer", "Schedule Call Back (Call Back)"}

# DXB team roster: owner_id -> display name
DXB_OWNERS = {
    "425098599": "Jully Gill",
    "81129092":  "Akanksha Yadav",
    "89994865":  "Ambreen Syed",
    "79152876":  "Insha Khan",
    "76337310":  "Ahlam Khandoq",
    "76337312":  "Patrecia Haddad",
    "457296009": "Rahul",
    "77931703":  "Abhi V",
    "82756823":  "Arya",
    "594801542": "Wahab Saeed Dogar",
    "1186837974":"Asfandyar Malik",
    "78332276":  "Kawleen Kaur",
    "89398738":  "Komal Zahid",
    "82714205":  "Muhammad Jalal Shah",
    "86887642":  "Khurram Iqbal",
    "2111743372":"Ronalyn Aguilar",
    "331190104": "Aleen Naeem",
    "331190099": "Ayesha Anum",
    "331190101": "Madah Khan",
    "93415418":  "Sneha",
}
OWNER_IDS = list(DXB_OWNERS.keys())
PORTAL_ID = "23735726"

# ----------------------------------------------------------------------------
# HTTP helper with basic rate-limit handling
# ----------------------------------------------------------------------------
def post(url, payload, tries=5):
    for attempt in range(tries):
        r = requests.post(url, headers=HEADERS, json=payload)
        if r.status_code == 429:
            time.sleep(2 * (attempt + 1))
            continue
        if r.status_code >= 400:
            # surface HubSpot's actual error message, not a bare status code
            try:
                detail = r.json()
            except Exception:
                detail = r.text
            raise SystemExit(f"\nHubSpot {r.status_code} error on {url}\n"
                             f"Payload: {payload}\nResponse: {detail}\n")
        return r.json()
    r.raise_for_status()


def to_ms(date_str):
    return int(dt.datetime.strptime(date_str, "%Y-%m-%d")
              .replace(tzinfo=dt.timezone.utc).timestamp() * 1000)


def parse_ts(ts):
    """Accept either epoch-millis (int/str) or an ISO8601 string like
    '2026-04-12T17:18:37.382Z'. Returns a tz-aware datetime, or None."""
    s = str(ts)
    if s.isdigit():
        return dt.datetime.fromtimestamp(int(s) / 1000, tz=dt.timezone.utc)
    try:
        return dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


# ----------------------------------------------------------------------------
# STEP 1 — Pull all DXB contacts created in the window (the denominator)
# ----------------------------------------------------------------------------
def fetch_contacts():
    url = f"{BASE}/crm/v3/objects/contacts/search"
    contacts = {}
    after = None
    while True:
        payload = {
            "filterGroups": [{
                "filters": [
                    {"propertyName": "createdate", "operator": "GTE", "value": str(to_ms(WINDOW_START))},
                    {"propertyName": "createdate", "operator": "LT",  "value": str(to_ms(WINDOW_END))},
                    {"propertyName": "hubspot_owner_id", "operator": "IN", "values": OWNER_IDS},
                ]
            }],
            "properties": ["firstname", "lastname", "email", "phone", "createdate",
                           "lead_stage", "hubspot_owner_id"],
            "limit": 100,
        }
        if after:
            payload["after"] = after
        data = post(url, payload)
        for r in data.get("results", []):
            p = r["properties"]
            contacts[r["id"]] = {
                "id": r["id"],
                "firstname": p.get("firstname") or "",
                "lastname": p.get("lastname") or "",
                "email": p.get("email") or "",
                "phone": p.get("phone") or "",
                "createdate": (p.get("createdate") or "")[:10],
                "lead_stage": p.get("lead_stage") or "(none)",
                "owner_id": p.get("hubspot_owner_id") or "",
            }
        paging = data.get("paging", {}).get("next", {}).get("after")
        if not paging:
            break
        after = paging
    return contacts


# ----------------------------------------------------------------------------
# STEP 2 — Pull all Call engagements for DXB owners in the window,
#          with their timestamps and associated contact IDs.
# ----------------------------------------------------------------------------
def fetch_calls():
    """Page through calls PER OWNER to stay well under the 10,000-result search
    ceiling and avoid large IN+date filter combinations that HubSpot can reject."""
    url = f"{BASE}/crm/v3/objects/calls/search"
    calls_by_contact = defaultdict(list)

    for oid in OWNER_IDS:
        after = None
        while True:
            payload = {
                "filterGroups": [{
                    "filters": [
                        {"propertyName": "hs_timestamp", "operator": "GTE", "value": str(to_ms(CALL_START))},
                        {"propertyName": "hs_timestamp", "operator": "LT",  "value": str(to_ms(CALL_END))},
                        {"propertyName": "hubspot_owner_id", "operator": "EQ", "value": oid},
                    ]
                }],
                "sorts": [{"propertyName": "hs_timestamp", "direction": "ASCENDING"}],
                "properties": ["hs_timestamp", "hubspot_owner_id"],
                "limit": 100,
            }
            if after:
                payload["after"] = after
            data = post(url, payload)
            results = data.get("results", [])
            ids = [r["id"] for r in results]
            ts_map = {r["id"]: r["properties"].get("hs_timestamp") for r in results}

            if ids:
                assoc = post(f"{BASE}/crm/v4/associations/calls/contacts/batch/read",
                             {"inputs": [{"id": i} for i in ids]})
                for row in assoc.get("results", []):
                    call_id = row["from"]["id"]
                    ts = ts_map.get(call_id)
                    if not ts:
                        continue
                    when = parse_ts(ts)
                    if when is None:
                        continue
                    for to in row.get("to", []):
                        calls_by_contact[to["toObjectId"]].append(when)

            paging = data.get("paging", {}).get("next", {}).get("after")
            if not paging:
                break
            after = paging
    return calls_by_contact


# ----------------------------------------------------------------------------
# STEP 3 — Compute flags
# ----------------------------------------------------------------------------
def max_gap_days(times):
    if len(times) < 2:
        return 0
    s = sorted(times)
    return max((s[i + 1] - s[i]).days for i in range(len(s) - 1))


def analyse(contacts, calls_by_contact):
    rows = []
    for cid, c in contacts.items():
        if c["lead_stage"] in EXCLUDE_STAGES:
            continue
        times = calls_by_contact.get(int(cid), [])
        n = len(times)
        gap = max_gap_days(times)
        flag_few = n < MIN_CALLS
        flag_gap = gap >= GAP_DAYS
        if not (flag_few or flag_gap):
            continue
        flags = []
        if flag_few:
            flags.append(f"<{MIN_CALLS} calls")
        if flag_gap:
            flags.append(f"gap {gap}d")
        rows.append({
            **c,
            "owner_name": DXB_OWNERS.get(c["owner_id"], c["owner_id"]),
            "call_count": n,
            "max_gap": gap,
            "flag_few": flag_few,
            "flag_gap": flag_gap,
            "flags": " + ".join(flags),
            "call_dates": ", ".join(t.strftime("%Y-%m-%d") for t in sorted(times)),
        })
    # Most under-worked first: fewest calls, then biggest gap
    rows.sort(key=lambda r: (r["call_count"], -r["max_gap"]))
    return rows


# ----------------------------------------------------------------------------
# STEP 4 — Write Excel
# ----------------------------------------------------------------------------
def write_excel(rows, contacts, path="dxb_call_gap_analysis.xlsx"):
    wb = Workbook()

    # ---- Sheet 1: README ----
    ws0 = wb.active
    ws0.title = "README"
    ws0["A1"] = "DXB Team — Call-Attempt Gap Analysis"
    ws0["A1"].font = Font(name="Arial", bold=True, size=16, color="16205E")
    lines = [
        "",
        "Filter applied:",
        f"  - Contact created between {WINDOW_START} and {WINDOW_END}",
        f"  - Owner is one of the {len(DXB_OWNERS)} DXB team members",
        f"  - Lead stage is NOT 'No Answer' and NOT 'Schedule Call Back (Call Back)'",
        "",
        "A lead is flagged if EITHER:",
        f"  - It has fewer than {MIN_CALLS} logged Call activities, OR",
        f"  - There is a gap of {GAP_DAYS}+ days between two consecutive call attempts",
        "",
        "Only logged Call engagements are counted (not emails, WhatsApp, or notes).",
        "",
        f"Total DXB contacts in window: {len(contacts)}",
        f"Total flagged leads: {len(rows)}",
        f"Snapshot: {dt.date.today().isoformat()}",
        "",
        "Sheets:",
        "  - Flagged Leads — every lead that triggered a flag, with call count, max gap, and call dates",
        "  - By Owner — counts per owner",
    ]
    for i, ln in enumerate(lines, 2):
        ws0.cell(row=i, column=1, value=ln).font = Font(name="Arial", size=11)
    ws0.column_dimensions["A"].width = 100

    # ---- Sheet 2: Flagged Leads ----
    ws = wb.create_sheet("Flagged Leads")
    headers = ["#", "Owner", "Full Name", "Email", "Phone", "Created",
               "Lead Stage", "Call Attempts", "Max Gap (days)", "Flags",
               "Call Dates", "HubSpot Link"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", start_color="16205E")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    red = PatternFill("solid", start_color="F4CCCC")
    amber = PatternFill("solid", start_color="FFE599")
    for i, r in enumerate(rows, 1):
        full = (r["firstname"] + " " + r["lastname"]).strip() or "(no name)"
        link = f"https://app.hubspot.com/contacts/{PORTAL_ID}/contact/{r['id']}"
        ws.append([i, r["owner_name"], full, r["email"], r["phone"], r["createdate"],
                   r["lead_stage"], r["call_count"], r["max_gap"], r["flags"],
                   r["call_dates"], link])
        # color the call-count cell
        cc = ws.cell(row=i + 1, column=8)
        if r["call_count"] == 0:
            cc.fill = red
        elif r["call_count"] < MIN_CALLS:
            cc.fill = amber
        # color the gap cell
        gc = ws.cell(row=i + 1, column=9)
        if r["flag_gap"]:
            gc.fill = red
        lc = ws.cell(row=i + 1, column=12)
        lc.hyperlink = link
        lc.font = Font(name="Arial", color="0563C1", underline="single")

    widths = {1: 5, 2: 20, 3: 26, 4: 32, 5: 18, 6: 12, 7: 22,
              8: 13, 9: 14, 10: 18, 11: 50, 12: 46}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if not (cell.font and cell.font.underline):
                cell.font = Font(name="Arial", size=10)
    ws.freeze_panes = "A2"

    # ---- Sheet 3: By Owner ----
    ws2 = wb.create_sheet("By Owner")
    ws2.append(["Owner", "Flagged Leads", "Avg Calls", "Leads with 0 Calls"])
    for c in range(1, 5):
        cell = ws2.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    by_owner = defaultdict(list)
    for r in rows:
        by_owner[r["owner_name"]].append(r)
    summary = []
    for owner, items in by_owner.items():
        avg = sum(x["call_count"] for x in items) / len(items)
        zeros = sum(1 for x in items if x["call_count"] == 0)
        summary.append((owner, len(items), round(avg, 1), zeros))
    summary.sort(key=lambda x: -x[1])
    for owner, cnt, avg, zeros in summary:
        ws2.append([owner, cnt, avg, zeros])
    for col, w in {"A": 22, "B": 15, "C": 12, "D": 18}.items():
        ws2.column_dimensions[col].width = w
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
    ws2.freeze_panes = "A2"

    wb.save(path)
    print(f"Saved {path} — {len(rows)} flagged leads out of {len(contacts)} contacts.")


# ----------------------------------------------------------------------------
if __name__ == "__main__":
    print("1/3  Fetching DXB contacts created in window ...")
    contacts = fetch_contacts()
    print(f"     {len(contacts)} contacts.")

    print("2/3  Fetching call engagements + associations ...")
    calls_by_contact = fetch_calls()
    print(f"     calls tied to {len(calls_by_contact)} contacts.")

    print("3/3  Computing flags and writing Excel ...")
    rows = analyse(contacts, calls_by_contact)
    write_excel(rows, contacts)
